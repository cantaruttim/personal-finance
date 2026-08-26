import re
import pandas as pd
from pathlib import Path
import pdfplumber
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# =============================================================================
# 1. FUNÇÕES AUXILIARES
# =============================================================================

def parse_currency(val):
    """Converte string no formato brasileiro (1.234,56) para float."""
    if not val:
        return None
    val = str(val).strip().replace('R$', '').replace(' ', '')
    try:
        # Remove pontos de milhar e troca vírgula por ponto
        return float(val.replace('.', '').replace(',', '.'))
    except:
        return None

def classify_item(code, desc):
    """Classifica um item como 'Provento' ou 'Desconto' baseado no código e descrição."""
    # Códigos conhecidos (sem zeros à esquerda)
    provento_codes = ['00010', '00020', '02000', '02005', '02045', '10260']
    desconto_codes = ['5000', '5030', '5100', '5710', '5715', '6195', '6600', '8145', '6705']
    
    if code in provento_codes:
        return 'Provento'
    if code in desconto_codes:
        return 'Desconto'
    
    # Fallback por palavra-chave
    desc_upper = desc.upper()
    provento_keywords = ['SALARIO BASE', 'GRATIFICACAO', 'FERIAS', 'PROVISAO FERIAS', 'SALARIO NOMINAL']
    desconto_keywords = ['INSS', 'IRRF', 'DESC ADTO', 'SEGURO DE VIDA', 'ASSIST MEDICA',
                         'VGBL', 'COPARTICIPAÇÃO', 'PARCELA ADIANTAMENTO']
    
    for kw in provento_keywords:
        if kw in desc_upper:
            return 'Provento'
    for kw in desconto_keywords:
        if kw in desc_upper:
            return 'Desconto'
    
    # Se não encaixar, assume desconto (segurança)
    return 'Desconto'

def calcular_inss(salario_bruto):
    """Calcula INSS conforme tabela 2025 (ajustar para 2026 quando disponível)."""
    if not salario_bruto or salario_bruto <= 0:
        return 0
    faixas = [
        (1412.00, 0.075),
        (2666.68, 0.09),
        (4000.03, 0.12),
        (9101.06, 0.14)
    ]
    total = 0
    base = salario_bruto
    for i, (limite, aliquota) in enumerate(faixas):
        if base <= 0:
            break
        if i == 0:
            teto = limite
        else:
            teto = limite - faixas[i-1][0]
        if base > teto:
            total += teto * aliquota
            base -= teto
        else:
            total += base * aliquota
            base = 0
    return total

def calcular_irrf(base_irpf):
    """Calcula IRRF simplificado (sem deduções) para estimativa."""
    if not base_irpf or base_irpf <= 0:
        return 0
    faixas = [
        (2259.20, 0.0),
        (2826.65, 0.075),
        (3751.05, 0.15),
        (4664.68, 0.225),
        (float('inf'), 0.275)
    ]
    deducoes = [0, 169.44, 381.44, 662.77, 896.00]
    for i, (limite, aliquota) in enumerate(faixas):
        if base_irpf <= limite:
            return max(0, base_irpf * aliquota - deducoes[i])
    return max(0, base_irpf * 0.275 - 896.00)

# =============================================================================
# 2. EXTRAÇÃO DO PDF (NOVA ABORDAGEM - LINHA A LINHA)
# =============================================================================

def parse_pdf(pdf_path):
    """Extrai informações do holerite usando análise linha a linha do texto."""
    with pdfplumber.open(pdf_path) as pdf:
        page = pdf.pages[0]
        text = page.extract_text() or ''
        
        # ----- Metadados (via regex no texto) -----
        def get_pattern(pattern):
            match = re.search(pattern, text, re.IGNORECASE)
            return match.group(1).strip() if match else None

        referencia = get_pattern(r'Referência:\s*(\d{2}/\d{4})')
        periodo = get_pattern(r'Período de Pagamento:\s*([\d/]+) a ([\d/]+)')
        nome = get_pattern(r'Nome:\s*([^\n]+?)(?=Identificador|CPF|$)')
        cpf = get_pattern(r'CPF:\s*([\d\.]+-[\d]{2})')
        cargo = get_pattern(r'Cargo:\s*([^\n]+?)(?=Tipo de salário|$)')
        salario_contratual = get_pattern(r'Salário contratual:\s*([\d\.]+,[\d]{2})')
        total_proventos = get_pattern(r'Total de proventos:\s*([\d\.]+,[\d]{2})')
        total_descontos = get_pattern(r'Total de descontos:\s*([\d\.]+,[\d]{2})')
        total_liquido = get_pattern(r'Total líquido:\s*([\d\.]+,[\d]{2})')
        base_inss = get_pattern(r'Sal\. contrib\. INSS:\s*([\d\.]+,[\d]{2})')
        base_fgts = get_pattern(r'Base cal\. FGTS:\s*([\d\.]+,[\d]{2})')
        fgts_mes = get_pattern(r'FGTS mês:\s*([\d\.]+,[\d]{2})')
        base_irpf = get_pattern(r'Base calc\. IRPF:\s*([\d\.]+,[\d]{2})')

        salario_contratual = parse_currency(salario_contratual)
        total_proventos = parse_currency(total_proventos)
        total_descontos = parse_currency(total_descontos)
        total_liquido = parse_currency(total_liquido)
        base_inss = parse_currency(base_inss)
        base_fgts = parse_currency(base_fgts)
        fgts_mes = parse_currency(fgts_mes)
        base_irpf = parse_currency(base_irpf)

        # ----- Extração de itens (proventos e descontos) linha a linha -----
        items = []
        # Divide o texto em linhas
        lines = text.split('\n')
        for line in lines:
            line = line.strip()
            # Procura por linha que começa com código de 5 dígitos (ex: 00010, 5000)
            # O código pode ter 5 dígitos, mas alguns como 5000 são 4 dígitos, então usamos 4 ou 5
            match_code = re.match(r'^(\d{4,5})', line)
            if not match_code:
                continue
            code = match_code.group(1)
            # Remove o código do início da linha
            rest = line[len(code):]
            
            # Agora precisamos extrair a descrição e os valores.
            # Padrão: descrição seguida de números (qtde, provento, desconto, resultado)
            # Vamos procurar por números no formato brasileiro na linha
            numbers = re.findall(r'(\d{1,3}(?:\.\d{3})*,\d{2})', rest)
            if len(numbers) < 2:
                # Se não houver números, pula
                continue
            
            # O primeiro número é a quantidade (qtde), o segundo é provento, o terceiro desconto, o quarto resultado
            # Mas alguns itens têm apenas provento ou apenas desconto.
            # Vamos pegar a descrição como o texto entre o código e o primeiro número
            # Usamos a posição do primeiro número
            first_num_pos = re.search(r'\d{1,3}(?:\.\d{3})*,\d{2}', rest)
            if not first_num_pos:
                continue
            desc = rest[:first_num_pos.start()].strip()
            
            # Converte os números
            qtde = parse_currency(numbers[0]) if len(numbers) > 0 else 0
            provento = parse_currency(numbers[1]) if len(numbers) > 1 else 0
            desconto = parse_currency(numbers[2]) if len(numbers) > 2 else 0
            # resultado = parse_currency(numbers[3]) if len(numbers) > 3 else 0 (não precisamos)
            
            # Se provento > 0, adiciona como provento
            if provento and provento > 0:
                items.append({
                    'code': code,
                    'description': desc,
                    'amount': provento,
                    'tipo': 'Provento'
                })
            # Se desconto > 0, adiciona como desconto
            if desconto and desconto > 0:
                items.append({
                    'code': code,
                    'description': desc,
                    'amount': desconto,
                    'tipo': 'Desconto'
                })
            # Se ambos forem zero, pode ser que a linha seja apenas qtde e resultado, mas isso não ocorre
        
        # Se ainda assim não encontrou itens, tenta um fallback com regex mais amplo
        if not items:
            # Última tentativa: procurar padrões com código, descrição, e valores
            pattern = re.compile(
                r'(\d{4,5})'           # código (4 ou 5 dígitos)
                r'([A-Za-zÀ-ÿ\s/&]+?)' # descrição (letras, espaços, símbolos)
                r'(\d{1,3}(?:\.\d{3})*,\d{2})' # qtde
                r'(\d{1,3}(?:\.\d{3})*,\d{2})' # provento
                r'(\d{1,3}(?:\.\d{3})*,\d{2})' # desconto
                r'(\d{1,3}(?:\.\d{3})*,\d{2})' # resultado
                r'(?=\d{4,5}|$)',
                re.DOTALL
            )
            for match in pattern.finditer(text):
                code = match.group(1)
                desc = match.group(2).strip()
                provento = parse_currency(match.group(4))
                desconto = parse_currency(match.group(5))
                if provento and provento > 0:
                    items.append({'code': code, 'description': desc, 'amount': provento, 'tipo': 'Provento'})
                if desconto and desconto > 0:
                    items.append({'code': code, 'description': desc, 'amount': desconto, 'tipo': 'Desconto'})

        # Retorna o resultado
        return {
            'referencia': referencia,
            'periodo_inicio': periodo[0] if periodo else None,
            'periodo_fim': periodo[1] if periodo else None,
            'nome': nome,
            'cpf': cpf,
            'cargo': cargo,
            'salario_contratual': salario_contratual,
            'total_proventos': total_proventos,
            'total_descontos': total_descontos,
            'total_liquido': total_liquido,
            'base_inss': base_inss,
            'base_fgts': base_fgts,
            'fgts_mes': fgts_mes,
            'base_irpf': base_irpf,
            'items': items
        }

# =============================================================================
# 3. PROCESSAMENTO
# =============================================================================

pasta_holerites = Path('data/holerites/matheus/2026')
if not pasta_holerites.exists():
    print(f'ERRO: Pasta não encontrada -> {pasta_holerites.absolute()}')
    exit()

pdf_files = sorted(pasta_holerites.glob('Recibo Pagto *.pdf'))
if not pdf_files:
    print(f'Nenhum PDF encontrado em {pasta_holerites}')
    exit()

print(f'Encontrados {len(pdf_files)} holerites. Processando...')

all_months = []
all_items = []
salarios_contratuais = []

for pdf_path in pdf_files:
    print(f'  Lendo: {pdf_path.name}')
    data = parse_pdf(pdf_path)
    
    if not data['referencia']:
        ref_match = re.search(r'(\d{2})-(\d{4})', pdf_path.name)
        data['referencia'] = f"{ref_match.group(1)}/{ref_match.group(2)}" if ref_match else pdf_path.stem

    if data['salario_contratual']:
        salarios_contratuais.append(data['salario_contratual'])

    # Resumo mensal
    month_row = {
        'Mês/Ano': data['referencia'],
        'Salário Contratual': data['salario_contratual'],
        'Total Proventos': data['total_proventos'],
        'Total Descontos': data['total_descontos'],
        'Total Líquido': data['total_liquido'],
        'Base INSS': data['base_inss'],
        'Base FGTS': data['base_fgts'],
        'FGTS mês': data['fgts_mes'],
        'Base IRPF': data['base_irpf'],
        'Nome': data['nome'],
        'CPF': data['cpf'],
        'Cargo': data['cargo']
    }
    all_months.append(month_row)

    # Itens detalhados
    for item in data['items']:
        all_items.append({
            'Mês/Ano': data['referencia'],
            'Código': item['code'],
            'Descrição': item['description'],
            'Tipo': item['tipo'],
            'Valor': item['amount']
        })

# =============================================================================
# 4. DATAFRAMES
# =============================================================================

df_summary = pd.DataFrame(all_months).sort_values('Mês/Ano').reset_index(drop=True)
cols_summary = ['Mês/Ano', 'Salário Contratual', 'Total Proventos', 'Total Descontos',
                'Total Líquido', 'Base INSS', 'Base FGTS', 'FGTS mês', 'Base IRPF',
                'Nome', 'CPF', 'Cargo']
df_summary = df_summary[cols_summary]

if all_items:
    df_items = pd.DataFrame(all_items)
    df_items_pivot = df_items.pivot_table(
        index=['Mês/Ano', 'Código', 'Descrição'],
        columns='Tipo',
        values='Valor',
        aggfunc='sum'
    ).reset_index().rename_axis(None, axis=1)
    for col in ['Provento', 'Desconto']:
        if col not in df_items_pivot.columns:
            df_items_pivot[col] = 0.0
    df_items_pivot.fillna(0, inplace=True)
    df_items_pivot = df_items_pivot.sort_values(['Mês/Ano', 'Código']).reset_index(drop=True)
    df_items_pivot = df_items_pivot[['Mês/Ano', 'Código', 'Descrição', 'Provento', 'Desconto']]
else:
    df_items_pivot = pd.DataFrame(columns=['Mês/Ano', 'Código', 'Descrição', 'Provento', 'Desconto'])

# =============================================================================
# 5. CÁLCULO DO 13º
# =============================================================================

decimo_info = {}
if salarios_contratuais:
    media_salarial = sum(salarios_contratuais) / len(salarios_contratuais)
    decimo_bruto = media_salarial
    inss_13 = calcular_inss(decimo_bruto)
    base_irrf_13 = decimo_bruto - inss_13
    irrf_13 = calcular_irrf(base_irrf_13) if base_irrf_13 > 0 else 0
    decimo_liquido = decimo_bruto - inss_13 - irrf_13
    primeira_parcela = decimo_bruto / 2
    segunda_parcela = decimo_liquido - primeira_parcela
    
    decimo_info = {
        'Salário médio (base)': media_salarial,
        '13º Bruto Estimado': decimo_bruto,
        'INSS sobre 13º': inss_13,
        'IRRF sobre 13º': irrf_13,
        '13º Líquido Estimado': decimo_liquido,
        '1ª Parcela (até 30/11)': primeira_parcela,
        '2ª Parcela (estimada)': segunda_parcela
    }

# =============================================================================
# 6. EXPORTAÇÃO EXCEL
# =============================================================================

output_file = 'Holerites_Consolidado.xlsx'
with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
    df_summary.to_excel(writer, sheet_name='Resumo Mensal', index=False)
    df_items_pivot.to_excel(writer, sheet_name='Detalhamento', index=False)
    if decimo_info:
        df_decimo = pd.DataFrame([decimo_info])
        df_decimo.to_excel(writer, sheet_name='13º Salário (estimado)', index=False)

# =============================================================================
# 7. FORMATAÇÃO
# =============================================================================

wb = load_workbook(output_file)
thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                     top=Side(style='thin'), bottom=Side(style='thin'))
header_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
header_font = Font(bold=True)

def format_sheet(sheet, currency_cols=None):
    for col in sheet.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                max_len = max(max_len, len(str(cell.value or '')))
            except:
                pass
        sheet.column_dimensions[col_letter].width = min(max_len + 2, 50)
    
    for row in sheet.iter_rows():
        for cell in row:
            cell.border = thin_border
            if cell.row == 1:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center')
    
    if currency_cols:
        for col_name in currency_cols:
            col_letter = None
            for cell in sheet[1]:
                if cell.value == col_name:
                    col_letter = get_column_letter(cell.column)
                    break
            if col_letter:
                for row in range(2, sheet.max_row + 1):
                    cell = sheet[f"{col_letter}{row}"]
                    if cell.value is not None:
                        cell.number_format = '#,##0.00_-'

format_sheet(wb['Resumo Mensal'], currency_cols=[
    'Salário Contratual', 'Total Proventos', 'Total Descontos', 'Total Líquido',
    'Base INSS', 'Base FGTS', 'FGTS mês', 'Base IRPF'
])

format_sheet(wb['Detalhamento'], currency_cols=['Provento', 'Desconto'])

if decimo_info:
    format_sheet(wb['13º Salário (estimado)'], currency_cols=list(decimo_info.keys()))

wb.save(output_file)

print('\n' + '=' * 60)
print(f'✅ PLANILHA GERADA COM SUCESSO!')
print(f'📁 Arquivo: {output_file}')
print(f'📊 Resumo: {len(df_summary)} meses')
print(f'📋 Detalhamento: {len(df_items_pivot)} itens')
if decimo_info:
    print(f'🎄 13º Bruto Estimado: R$ {decimo_info["13º Bruto Estimado"]:.2f}')
    print(f'🎄 13º Líquido Estimado: R$ {decimo_info["13º Líquido Estimado"]:.2f}')
print('=' * 60)